from pathlib import Path
import pandas as pd
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
os.system("clear")  # macOS/Linux

#LOADING OF FILES#
excel_path = Path(__file__).resolve().parent / "grades.xlsx"
grades = pd.read_excel(excel_path)


#STUDENT LEVEL CALCULATION(QUIZ AVERAGE, FINAL SCORE AND RESULT #)
grades['QuizAverage'] = grades[['Quiz1','Quiz2']].mean(axis=1)
grades['FinalScore'] = (grades['QuizAverage'] * .30) + (grades['Attendance'] * .20 + (grades['Exam']* .50))
grades['Result'] = grades['FinalScore'].apply(lambda x: "Pass" if x >= 80 else "Fail")
display_columns =[
    'StudentID',
    'StudentName',
    'Subject',
    'Quiz1',
    'Quiz2',
    'Attendance',
    'QuizAverage',
    'FinalScore',
    'Result',
    'Grade Band',
    'Risk Level'
]

def grade_band(score):
    if score >= 90:
        return "A"
    elif score >= 80:
        return "B"
    elif score >= 70:
        return "C"
    else:
        return "D/F"

def risk_level(row):
    if row["FinalScore"] < 75:
        return "High Risk"
    elif row["Attendance"] < 85:
        return "Attendance Risk"
    elif row["FinalScore"] < 80:
        return "Watchlist"
    else:
        return "On Track"
grades["Grade Band"] = grades["FinalScore"].apply(grade_band)
grades["Risk Level"] = grades.apply(risk_level,axis=1)


#SUMMARY METRICS#
final_students =grades['StudentID'].nunique()
final_students_per_subject =grades.groupby("Subject")["StudentID"].nunique()
min_final_score =grades["FinalScore"].min().round(2)
Overall_average_final_score = grades["FinalScore"].mean().round(2)
median_final_score = grades["FinalScore"].median().round(2)
max_final_score = grades["FinalScore"].max().round(2)
pass_rate = ((grades["Result"] == 'Pass').sum()/final_students)*100

summary_metrics =pd.DataFrame({
    "Metric": ["Total Students","Average Final Score","Median Final Score","Highest Final Score","Lowest Final Score","Pass Rate"],
    "Value": [final_students,Overall_average_final_score,median_final_score,max_final_score,min_final_score,pass_rate]
})
# print(summary_metrics)

#DATA VALIDATION#

#Check for duplicate records
duplicates = grades.duplicated(subset='StudentID').sum()
duplicate_info = grades.drop_duplicates()
print(duplicate_info)
#Data types validation
numeric_columns = ['Quiz1','Quiz2','Exam','Attendance']

#Invalid or Zero Grade Values

zero_grade_values = grades[numeric_columns].eq(0).sum()
invalid_quiz1_scores_count = (grades["Quiz1"].isna()| grades["Quiz1"].eq("")).sum()
invalid_quiz2_score_count = (grades["Quiz2"].isna()| grades["Quiz2"].eq("")).sum()
invalid_exam_score_count = (grades["Exam"].isna().sum()| grades["Exam"].eq("").sum())
invalid_attendance_score_count = (grades["Attendance"].isna().sum()| grades["Attendance"].eq("").sum())


validation_report_summary = pd.DataFrame({
    "Validation Check": ["Duplicate Student IDs", "Invalid Quiz1 Scores","Invalid Quiz2 Scores","Invalid Exam Scores","Invalid Attendance Scores"],
    "Result": [duplicates,invalid_quiz1_scores_count,invalid_quiz2_score_count,invalid_exam_score_count,invalid_attendance_score_count]
})
# print(validation_report_summary)



# #SUBJECT LEVEL ANALYSIS#
students_per_subject = grades.groupby('Subject')['StudentID'].count()

grouped  = grades.groupby('Subject')
overall_average_quiz_score = grades.groupby("Subject")[["Quiz1","Quiz2"]].mean().mean(axis=1).round(2)
overall_average_attendance = grades.groupby("Subject")["Attendance"].mean().round(2)
overall_average_exam_score = grades.groupby("Subject")["Exam"].mean().round(2)
overall_average_finalScore = grades.groupby("Subject")["FinalScore"].mean().round(2)
highest_final_score = grades.groupby("Subject")["FinalScore"].max()
lowest_final_score = grades.groupby("Subject")["FinalScore"].min()
pass_count = grades.groupby("Subject")["Result"].apply(lambda result:(result =="Pass").sum())
fail_count = grades.groupby("Subject")["Result"].apply(lambda result: (result =="Fail").sum())
pass_rate = grades.groupby("Subject")["Result"].apply(lambda result: (result =="Pass").mean()*100).round(2)

subject_analysis =pd.DataFrame({
    "StudentCount":students_per_subject,
    "AverageQuizScore":overall_average_quiz_score,
    "AverageExamScore":Overall_average_final_score,
    "AverageAttendance":overall_average_attendance,
    "AverageFinalScore":Overall_average_final_score,
    "HighestFinalScore":highest_final_score,
    "LowestFinalScore":lowest_final_score,
    "PassCount":pass_count,
    "FailCount":fail_count,
    "PassRate":pass_rate

}).reset_index()
# print(subject_analysis)


output_path = Path(__file__).resolve().parent / "teacher_grade_report.xlsx"

with pd.ExcelWriter(output_path) as writer:
    summary_metrics.to_excel(writer, sheet_name="Dashboard", index=False)
    grades.to_excel(writer, sheet_name="Student Results", index=False)
    subject_analysis.to_excel(writer, sheet_name="Subject Analysis", index=False)
    validation_report_summary.to_excel(writer, sheet_name="Validation Report", index=False)

print(f"Excel report created: {output_path}")
# Saving Charts into Excel File #


# VISUALS /CHARTS #
# - Total Students
ab = final_students_per_subject.plot(
    kind='bar',
    title="Total No. Of Students per Subject",
    legend=False,
    figsize=(8,4)

)
ab.set_xlabel("")  
for container in ab.containers:
    ab.bar_label(
        container,
        label_type='center',
        color='white',
        fontweight='bold',
    )
plt.xticks(rotation=0, ha="right")
plt.tight_layout()
no_of_students_path = Path(__file__).resolve().parent/"total_no_of_students.png"
plt.savefig(no_of_students_path, dpi=150, bbox_inches="tight")
plt.show()

# - Average Final Score per subject
average_final_score_per_subject = grades.groupby('Subject')['FinalScore'].mean().round(2)
print(average_final_score_per_subject)
ac = average_final_score_per_subject.plot(
    kind='bar',
    title='Average Final Score Per Subject',
    legend=False,
    figsize=(8,4)
)
ac.set_xlabel("")
ac.set_ylabel("Average Final Score")

for container in ac.containers:
    ac.bar_label(
        container,
        label_type="center",
        fmt="%.2f%%",
        color="white",
        fontweight="bold"
    )

plt.xticks(rotation=0)
plt.tight_layout()
average_final_score_path = Path(__file__).resolve().parent / "average_final_score.png"
plt.savefig(average_final_score_path, dpi=150, bbox_inches="tight")

plt.show()
# - Pass Rate 

pass_percentage_per_subject = (
    grades.groupby("Subject")["Result"]
    .apply(lambda result: (result == "Pass").mean() * 100)
    .round(2)
)

ad = pass_percentage_per_subject.plot(
    kind="bar",
    title="Pass Rate per Subject",
    figsize=(8,4),
    legend=False
)
ad.set_xlabel("")
ad.set_ylabel("Pass Percentage")

for container in ad.containers:
    ad.bar_label(
        container,
        label_type="center",
        fmt="%.2f%%",
        color="white",
        fontweight="bold"
    )
plt.xticks(rotation=0)
plt.tight_layout()
pass_percentage_path = Path(__file__).resolve().parent / "pass_percentage.png"
plt.savefig(pass_percentage_path, dpi=150, bbox_inches="tight")
plt.show()

# - Highest Score
# - Lowest Score

finalScore_highandlow = grades.groupby('Subject').agg(
    {'FinalScore':['max','min']}
)

ae = finalScore_highandlow.plot(
    kind="bar",
    stacked=False,
    title="Highest and Lowest Score per Subject",
    legend=False,
    figsize=(8,4)
)
ae.set_xlabel("")
ae.set_ylabel("")

for container in ae.containers:
    ae  .bar_label(
        container,
        label_type="center",
        fmt="%.2f",
        color="white",
        fontweight="bold"
    )
plt.xticks(rotation=0)
plt.tight_layout()
high_and_low_score_path = Path(__file__).resolve().parent / "high_and_low_score.png"
plt.savefig(high_and_low_score_path, dpi=150, bbox_inches="tight")
plt.show()


# - Pass vs Fail chart

pass_fail_percentage = pd.crosstab(
    grades["Subject"],
    grades["Result"],
    normalize="index"
) * 100
pass_fail_percentage = pass_fail_percentage.round(2)
print(pass_fail_percentage)


af = pass_fail_percentage.plot(
    kind="bar",
    legend=False,
    title='Pass vs Fail per Subject',
    figsize=(8,4)
)
af.set_xlabel("")
af.set_ylabel("")

for container in af.containers:
    af.bar_label(
        container,
        label_type="center",
        fmt="%.2f%%",
        color="white",
        fontweight="bold"
    )
plt.xticks(rotation=0)
plt.tight_layout()
pass_vs_fail_path = Path(__file__).resolve().parent / "pass_vs_fail.png"
plt.savefig(pass_vs_fail_path, dpi=150, bbox_inches="tight")
plt.show()

workbook = load_workbook(output_path)

if "Charts" in workbook.sheetnames:
    del workbook["Charts"]

charts_sheet = workbook.create_sheet("Charts")

charts_sheet.add_image(Image(no_of_students_path), "A1")
charts_sheet.add_image(Image(average_final_score_path), "A25")
charts_sheet.add_image(Image(pass_percentage_path), "A49")
charts_sheet.add_image(Image(high_and_low_score_path), "A73")
charts_sheet.add_image(Image(pass_vs_fail_path), "A97")

workbook.save(output_path)