from pathlib import Path
import pandas as pd
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
os.system("clear")  # macOS/Linux

#LOADING OF RAW FILE/DATA#
raw_data_path = Path(__file__).resolve().parent /"small_business_sales_raw_data.xlsx"
sales = pd.read_excel(raw_data_path)
sales = sales.dropna(subset=["Order ID"])
no_of_rows = sales.rows()
print(no_of_rows)

#Sales Grouping#
sales["Order Month"] = pd.to_datetime(sales["Order Date"]).dt.to_period("M")
grouped_monthly_sales = sales.groupby("Order Month")
jan_2025_sales = grouped_monthly_sales.get_group(pd.Period("2025-01",freq="M"))
grouped_branch = sales.groupby(["Order Month", "Branch"])
grouped_agent = sales.groupby(["Order Month","Branch","Sales Rep"])

#DATA VALIDATION#
#Check for duplicate order ID's
no_of_duplicate_order_id = (sales["Order ID"].value_counts() > 1).sum()

# print(no_of_duplicate_order_id)

duplicate_id_info = (
    sales.dropna(subset=["Order ID"])[sales.dropna(subset=["Order ID"]).duplicated(subset=["Order ID"], keep=False)][["Order ID", "Customer ID", "Customer Name"]]
    .drop_duplicates()
)
# print(duplicate_id_info)

#Check for NA values'
missing_values = sales.drop(columns="Notes",errors="ignore").isna().sum().sum()

data_validation_summary= pd.DataFrame({
    "Metric": ["Duplicate Order ID","Duplicate ID Info","Missing Values"],
    "Values": [no_of_duplicate_order_id,duplicate_id_info,missing_values]
})
print(data_validation_summary)

#MONTHLY KPI's
monthly_sales = grouped_monthly_sales["Gross Sales"].sum()
monthly_net_sales = grouped_monthly_sales["Net Sales"].sum()
monthly_units_sold = grouped_monthly_sales["Quantity"].sum()
average_order_value = (grouped_monthly_sales["Gross Sales"].sum()/monthly_units_sold).round(2)
net_sales_ratio = (monthly_net_sales/monthly_sales).round(2)* 100
# print(net_sales_ratio)

summary_monthly_KPI = pd.DataFrame({
     "Monthly Gross Sales": monthly_sales,
     "Monthly Net Sales": monthly_net_sales,
     "Monthly Units Sold": monthly_units_sold,
     "Average Order Value": average_order_value,
     "Net Sales Ratio":net_sales_ratio,
}).reset_index()
# print(summary_monthly_KPI)

#MONTHLY REVENUE BREAKDOWN

monthly_revenues_per_branch = grouped_branch["Gross Sales"].sum()
branch_revenue_share = (monthly_revenues_per_branch/monthly_sales).round(2) * 100
monthly_agent_revenue =grouped_agent["Gross Sales"].sum()
agent_revenue_share =(monthly_agent_revenue/monthly_sales).round(2) * 100


monthly_branch_revenue_breakdown = pd.DataFrame({
     "Monthly Branch Revenue": monthly_revenues_per_branch,
     "Branch Revenue Share": branch_revenue_share,
}).reset_index()
monthly_agent_revenue_breakdown = pd.DataFrame({
     "Monthly Agent Revenue": monthly_agent_revenue,
     "Agent Revenue Share":agent_revenue_share,
}).reset_index()



# #EXPORTING INTO EXCEL
output_path = Path(__file__).resolve().parent / "sales_analyzer.xlsx"
with pd.ExcelWriter(output_path) as writer:
    summary_monthly_KPI.to_excel(writer, sheet_name="Summary Monthly KPI", index=False)
    sales.to_excel(writer, sheet_name="Sales Raw Data",index=False)
    monthly_branch_revenue_breakdown .to_excel(writer, sheet_name="Branch Revenue Breakdown", index=False)
    monthly_agent_revenue_breakdown.to_excel(writer,sheet_name=" Agent Revenue Breakdown", index=False)
    data_validation_summary.to_excel(writer, sheet_name="Validation Report", index=False)

print(f"Excel report created: {output_path}")

# VISUALIZATION 


visual1 = monthly_sales.plot(
    kind="line",
    title="Monthly Gross Sales",
    legend=False,
    figsize=(8,4),
    grid=True,
    marker="s",
    linewidth = 3,
    rot=0
)
plt.xticks(rotation=0)
plt.tight_layout()
monthly_sales_fig_path = Path(__file__).resolve().parent/"monthly sales fig.png"
plt.savefig(monthly_sales_fig_path, dpi=150, bbox_inches="tight")
plt.close()


visual2 = monthly_units_sold.plot(
    kind="bar",
    title="Monthly Units Sold",
    legend=False,
    figsize=(11, 5),
    width=0.75,
    color="#4E79A7"
)

visual2.set_xlabel("")
visual2.set_ylabel("Units Sold")

for container in visual2.containers:
    labels = [f"{bar.get_height():,.0f}" for bar in container]

    visual2.bar_label(
        container,
        labels=labels,
        label_type="edge",
        padding=5,
        color="black",
        fontsize=9,
        fontweight="bold",
        bbox={
            "facecolor": "white",
            "edgecolor": "none",
            "boxstyle": "round,pad=0.2",
            "alpha": 0.85
        }
    )

visual2.margins(y=0.18)

plt.xticks(rotation=45, ha="right")
plt.tight_layout()
monthly_units_sold_fig_path = Path(__file__).resolve().parent/"monthly units sold fig.png"
plt.savefig(monthly_units_sold_fig_path, dpi=150, bbox_inches="tight")
plt.close()

# Monthly revenue per branch chart
branch_revenue_chart_data = monthly_revenues_per_branch.unstack()

visual3 = branch_revenue_chart_data.plot(
    kind="bar",
    figsize=(12, 5),
    width=0.8,
    title="Monthly Revenue per Branch"
)

visual3.set_xlabel("")
visual3.set_ylabel("Gross Sales")
visual3.legend(title="Branch", bbox_to_anchor=(1.02, 1), loc="upper left")

plt.xticks(rotation=45, ha="right")
plt.tight_layout()
monthly_branch_revenue_fig_path = Path(__file__).resolve().parent / "monthly_branch_revenue_fig.png"
plt.savefig(monthly_branch_revenue_fig_path, dpi=150, bbox_inches="tight")
plt.close()


#CHARTS EXPORTS

workbook = load_workbook(output_path)

if "Charts" in workbook.sheetnames:
    del workbook["Charts"]

charts_sheet = workbook.create_sheet("Charts")

charts_sheet.add_image(Image(monthly_sales_fig_path), "A1")
charts_sheet.add_image(Image(monthly_units_sold_fig_path ), "A25")
charts_sheet.add_image(Image(monthly_branch_revenue_fig_path), "A49")

workbook.save(output_path)